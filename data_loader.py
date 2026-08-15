from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from datetime import datetime, timezone
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd


WORKBOOK_BASENAME = "2026年天猫日报表-易丽洁8月"
DATE_SHEET_PATTERN = re.compile(r"^(\d{1,2})\.(\d{1,2})$")
SUPPORTED_EXTENSIONS = (".xls", ".xlsx", ".xlsm")
STORE_FILE_PATTERNS = {
    "易丽洁": ("2026年天猫日报表-易丽洁*",),
    "坐拥宁静": ("2026年坐拥宁静日报表*",),
    "国货严选": ("2026年国货严选日报表*",),
    "咖时光": ("2026年咖时光日报表-天猫*",),
}
LEGACY_SUMMARY_PRODUCT_ID = "全店汇总（早期格式）"
MIN_REPORT_DATE = pd.Timestamp(2026, 7, 1)
DISK_CACHE_SCHEMA_VERSION = "product-daily-v4"


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: list[list[object]]
    merged_ranges: list[tuple[int, int, int, int]]  # zero-based, end-exclusive


def _workbook_search_dirs(project_dir: Path) -> list[Path]:
    """Return workbook locations, with an explicit server data directory first."""
    directories: list[Path] = []
    configured = os.environ.get("TMALL_DATA_DIR")
    if configured:
        directories.append(Path(configured).expanduser())
    directories.extend([Path.home() / "Desktop", project_dir / "data", Path.cwd()])

    # Keep precedence while avoiding duplicate work when cwd equals project_dir.
    unique: list[Path] = []
    seen: set[str] = set()
    for directory in directories:
        key = str(directory.resolve())
        if key not in seen:
            seen.add(key)
            unique.append(directory)
    return unique


def find_workbook(explicit_path: str | Path | None = None) -> Path:
    """Find the daily workbook, preferring the user's Desktop source file."""
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.is_file():
            return path.resolve()
        raise FileNotFoundError(f"未找到指定 Excel 文件：{path}")

    project_dir = Path(__file__).resolve().parent
    search_dirs = _workbook_search_dirs(project_dir)
    candidates: list[Path] = []
    for directory in search_dirs:
        for suffix in SUPPORTED_EXTENSIONS:
            candidate = directory / f"{WORKBOOK_BASENAME}{suffix}"
            if candidate.is_file():
                candidates.append(candidate)
        if candidates:
            # The first directory with matches wins. Within it, prefer the source .xls.
            return sorted(candidates, key=lambda p: SUPPORTED_EXTENSIONS.index(p.suffix.lower()))[0].resolve()

    raise FileNotFoundError(
        f"未找到 {WORKBOOK_BASENAME}（支持 {', '.join(SUPPORTED_EXTENSIONS)}）。"
    )


def find_store_workbooks(
    explicit_paths: dict[str, str | Path] | None = None,
) -> dict[str, Path]:
    """Find one current workbook per store, preferring files on the Desktop.

    If multiple matching files exist in the same directory, the most recently
    modified supported workbook wins. This makes periodic file replacement or
    renamed monthly copies work without changing the dashboard code.
    """
    explicit_paths = explicit_paths or {}
    project_dir = Path(__file__).resolve().parent
    search_dirs = _workbook_search_dirs(project_dir)
    found: dict[str, Path] = {}
    missing: list[str] = []

    for store, patterns in STORE_FILE_PATTERNS.items():
        explicit = explicit_paths.get(store)
        if explicit:
            path = Path(explicit).expanduser()
            if not path.is_file():
                raise FileNotFoundError(f"{store}：未找到指定 Excel 文件：{path}")
            found[store] = path.resolve()
            continue

        selected: Path | None = None
        for directory in search_dirs:
            candidates = [
                path
                for pattern in patterns
                for path in directory.glob(pattern)
                if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
            ]
            if candidates:
                selected = max(candidates, key=lambda path: (path.stat().st_mtime_ns, path.name))
                break
        if selected is None:
            missing.append(store)
        else:
            found[store] = selected.resolve()

    if missing:
        raise FileNotFoundError(f"未找到以下店铺日报：{'、'.join(missing)}")
    return found


def load_store_daily(workbooks: dict[str, str | Path]) -> pd.DataFrame:
    """Load all stores while keeping store identity in every aggregate row."""
    frames: list[pd.DataFrame] = []
    for store, path_value in workbooks.items():
        path = Path(path_value)
        frame = load_product_daily_cached(path)
        cache_status = "磁盘缓存命中" if frame.attrs.get("cache_hit") else "磁盘缓存重建"
        frame.insert(0, "store", store)
        frame["source_file"] = str(path.resolve())
        frame["cache_status"] = cache_status
        frames.append(frame)
    if not frames:
        raise ValueError("没有可读取的店铺工作簿")
    return pd.concat(frames, ignore_index=True).sort_values(
        ["store", "date", "product_id"], ignore_index=True
    )


def _xls_sheets(path: Path) -> Iterable[SheetData]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("读取 .xls 需要 xlrd，请先执行 pip install -r requirements.txt") from exc

    book = xlrd.open_workbook(str(path), formatting_info=True)
    for sheet in book.sheets():
        rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
        yield SheetData(sheet.name, rows, list(sheet.merged_cells))


def _xlsx_sheets(path: Path) -> Iterable[SheetData]:
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True, read_only=False)
    try:
        for sheet in book.worksheets:
            rows = [
                [sheet.cell(row_idx, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                for row_idx in range(1, sheet.max_row + 1)
            ]
            merged_ranges = [
                (rng.min_row - 1, rng.max_row, rng.min_col - 1, rng.max_col)
                for rng in sheet.merged_cells.ranges
            ]
            yield SheetData(sheet.title, rows, merged_ranges)
    finally:
        book.close()


def _iter_sheets(path: Path) -> Iterable[SheetData]:
    if path.suffix.lower() == ".xls":
        yield from _xls_sheets(path)
    else:
        yield from _xlsx_sheets(path)


def _clean_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _find_header(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    required = {"商品ID", "数量", "单品结余"}
    for row_idx, row in enumerate(rows[:20]):
        mapping = {_clean_header(value): col_idx for col_idx, value in enumerate(row)}
        if required.issubset(mapping):
            return row_idx, mapping
    raise ValueError("前 20 行未找到包含 商品ID/数量/单品结余 的标题行")


def _find_legacy_header(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    """Find the old store-level layout that predates product IDs."""
    required = {"货号", "数量", "快递单量", "结余"}
    for row_idx, row in enumerate(rows[:20]):
        mapping = {_clean_header(value): col_idx for col_idx, value in enumerate(row)}
        if required.issubset(mapping):
            return row_idx, mapping
    raise ValueError("前 20 行未找到可识别的标准或早期日报标题行")


def _normalize_product_id(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        value = str(int(value)) if float(value).is_integer() else str(value)
    else:
        value = str(value).strip()
    value = value.removesuffix(".0")
    return value if re.fullmatch(r"\d{6,}", value) else None


def _number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    try:
        number = pd.to_numeric(value, errors="coerce")
        return 0.0 if pd.isna(number) else float(number)
    except (TypeError, ValueError):
        return 0.0


def _numeric_values(rows: list[list[object]], start_row: int, column: int) -> list[float]:
    values: list[float] = []
    for row in rows[start_row:]:
        if len(row) <= column or row[column] in (None, ""):
            continue
        value = pd.to_numeric(row[column], errors="coerce")
        if not pd.isna(value):
            values.append(float(value))
    return values


def _fractional_rate(values: list[float]) -> float:
    if not values:
        return 0.0
    return sum(abs(value - round(value)) > 1e-6 for value in values) / len(values)


def _correct_mislabeled_order_column(
    sheet: SheetData, header_idx: int, columns: dict[str, int], order_col: int | None
) -> int | None:
    """Correct a swapped promotion/order pair only when the data is unambiguous."""
    if order_col is None or order_col < 1:
        return order_col
    header_row = sheet.rows[header_idx]
    if len(header_row) <= order_col or _clean_header(header_row[order_col - 1]) != "推广花费":
        return order_col

    labelled_orders = _numeric_values(sheet.rows, header_idx + 1, order_col)
    adjacent_values = _numeric_values(sheet.rows, header_idx + 1, order_col - 1)
    if (
        len(labelled_orders) >= 5
        and len(adjacent_values) >= 5
        and _fractional_rate(labelled_orders) >= 0.25
        and _fractional_rate(adjacent_values) <= 0.05
    ):
        return order_col - 1
    return order_col


def _infer_year(path: Path, default: int = 2026) -> int:
    match = re.search(r"(20\d{2})", path.stem)
    return int(match.group(1)) if match else default


def _extract_sheet(sheet: SheetData, year: int) -> list[dict[str, object]]:
    date_match = DATE_SHEET_PATTERN.fullmatch(sheet.name.strip())
    if not date_match:
        return []
    sheet_date = pd.Timestamp(year, int(date_match.group(1)), int(date_match.group(2)))
    if sheet_date < MIN_REPORT_DATE:
        return []

    try:
        header_idx, columns = _find_header(sheet.rows)
    except ValueError:
        return _extract_legacy_sheet(sheet, year, date_match)
    product_col = columns["商品ID"]
    quantity_col = columns["数量"]
    order_col = _correct_mislabeled_order_column(
        sheet, header_idx, columns, columns.get("订单数")
    )
    profit_col = columns["单品结余"]
    sku_col = columns.get("货号")

    effective_ids: dict[int, str] = {}
    # Explicitly expand only merged product-ID regions. This is the key rule for
    # cases such as 8.13!A60:A63, whose quantities and M-column profit all belong
    # to the product ID stored in the top-left cell.
    for row_start, row_end, col_start, col_end in sheet.merged_ranges:
        if col_start <= product_col < col_end and row_end > header_idx + 1:
            top_row = max(row_start, header_idx + 1)
            raw_value = sheet.rows[row_start][product_col] if row_start < len(sheet.rows) else None
            product_id = _normalize_product_id(raw_value)
            if product_id:
                for row_idx in range(top_row, min(row_end, len(sheet.rows))):
                    effective_ids[row_idx] = product_id

    records: list[dict[str, object]] = []
    max_needed_col = max(product_col, quantity_col, profit_col, order_col or 0, sku_col or 0)
    for row_idx in range(header_idx + 1, len(sheet.rows)):
        row = sheet.rows[row_idx]
        if len(row) <= max_needed_col:
            continue
        product_id = effective_ids.get(row_idx) or _normalize_product_id(row[product_col])
        if not product_id:
            continue
        sku = "" if sku_col is None or row[sku_col] is None else str(row[sku_col]).strip()
        records.append(
            {
                "date": sheet_date,
                "sheet": sheet.name,
                "product_id": product_id,
                "sales_qty": _number(row[quantity_col]),
                "order_count": _number(row[order_col]) if order_col is not None else 0.0,
                "profit": _number(row[profit_col]),
                "sku": sku,
                "source_row": row_idx + 1,
            }
        )
    return records


def _extract_legacy_sheet(
    sheet: SheetData, year: int, date_match: re.Match[str]
) -> list[dict[str, object]]:
    """Preserve old sheets as one clearly labelled store-level aggregate.

    These sheets contain SKU/货号 detail quantities, but only one store-level
    快递单量 and 结余. Emitting all detail rows under a synthetic ID keeps the
    totals correct without falsely assigning the whole-store profit to one SKU.
    """
    header_idx, columns = _find_legacy_header(sheet.rows)
    sku_col = columns["货号"]
    quantity_col = columns["数量"]
    order_col = columns["快递单量"]
    profit_col = columns["结余"]
    max_needed_col = max(sku_col, quantity_col, order_col, profit_col)
    records: list[dict[str, object]] = []

    for row_idx in range(header_idx + 1, len(sheet.rows)):
        row = sheet.rows[row_idx]
        if len(row) <= max_needed_col:
            continue
        sku = "" if row[sku_col] is None else str(row[sku_col]).strip()
        if not sku or sku in {"合计", "总计", "小计"}:
            continue
        records.append(
            {
                "date": pd.Timestamp(year, int(date_match.group(1)), int(date_match.group(2))),
                "sheet": sheet.name,
                "product_id": LEGACY_SUMMARY_PRODUCT_ID,
                "sales_qty": _number(row[quantity_col]),
                "order_count": _number(row[order_col]),
                "profit": _number(row[profit_col]),
                "sku": sku,
                "source_row": row_idx + 1,
            }
        )
    return records


def load_product_daily(path: str | Path) -> pd.DataFrame:
    """Aggregate each date sheet by product ID across all covered detail rows."""
    workbook_path = Path(path)
    year = _infer_year(workbook_path)
    detail_records: list[dict[str, object]] = []
    for sheet in _iter_sheets(workbook_path):
        detail_records.extend(_extract_sheet(sheet, year))

    if not detail_records:
        raise ValueError("没有从日期 Sheet 中读取到商品数据")

    detail = pd.DataFrame(detail_records)
    daily = (
        detail.groupby(["date", "sheet", "product_id"], as_index=False)
        .agg(
            sales_qty=("sales_qty", "sum"),
            order_count=("order_count", "sum"),
            profit=("profit", "sum"),
            sku_count=("sku", lambda values: len({v for v in values if v})),
            skus=("sku", lambda values: "、".join(dict.fromkeys(v for v in values if v))),
            source_rows=("source_row", lambda values: ",".join(map(str, values))),
        )
        .sort_values(["date", "product_id"], ignore_index=True)
    )
    daily["sales_qty"] = daily["sales_qty"].round(6)
    daily["order_count"] = daily["order_count"].round(6)
    daily["profit"] = daily["profit"].round(6)
    return daily


def _workbook_fingerprint(path: Path) -> str:
    """Return a robust content fingerprint for automatic cache invalidation."""
    digest = hashlib.sha256()
    stat = path.stat()
    digest.update(DISK_CACHE_SCHEMA_VERSION.encode("utf-8"))
    digest.update(str(path.resolve()).encode("utf-8"))
    digest.update(str(stat.st_size).encode("ascii"))
    digest.update(str(stat.st_mtime_ns).encode("ascii"))
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _default_cache_dir() -> Path:
    configured = os.environ.get("TMALL_CACHE_DIR")
    return Path(configured) if configured else Path(__file__).resolve().parent / ".cache"


def load_product_daily_cached(
    path: str | Path, cache_dir: str | Path | None = None
) -> pd.DataFrame:
    """Load an aggregate from disk cache, rebuilding when the workbook changes.

    Cache files are derived data only. Writes are atomic and stale versions for
    the same source workbook are removed after a successful rebuild.
    """
    workbook_path = Path(path).resolve()
    target_dir = Path(cache_dir) if cache_dir else _default_cache_dir()
    target_dir.mkdir(parents=True, exist_ok=True)

    source_id = hashlib.sha256(str(workbook_path).encode("utf-8")).hexdigest()[:16]
    fingerprint = _workbook_fingerprint(workbook_path)
    cache_path = target_dir / f"{source_id}-{fingerprint}.pkl"
    metadata_path = target_dir / f"{source_id}-{fingerprint}.json"

    if cache_path.is_file():
        try:
            cached = pd.read_pickle(cache_path)
            required = {"date", "sheet", "product_id", "sales_qty", "order_count", "profit"}
            if required.issubset(cached.columns):
                cached.attrs["cache_hit"] = True
                cached.attrs["cache_path"] = str(cache_path)
                return cached
        except Exception:
            cache_path.unlink(missing_ok=True)
            metadata_path.unlink(missing_ok=True)

    daily = load_product_daily(workbook_path)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            dir=target_dir, prefix=f"{source_id}-", suffix=".tmp", delete=False
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        daily.to_pickle(temporary_path)
        os.replace(temporary_path, cache_path)
        metadata = {
            "schema_version": DISK_CACHE_SCHEMA_VERSION,
            "source": str(workbook_path),
            "fingerprint": fingerprint,
            "rows": len(daily),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink(missing_ok=True)

    for stale_path in target_dir.glob(f"{source_id}-*"):
        if stale_path not in {cache_path, metadata_path}:
            stale_path.unlink(missing_ok=True)

    daily.attrs["cache_hit"] = False
    daily.attrs["cache_path"] = str(cache_path)
    return daily


def complete_daily_series(daily: pd.DataFrame) -> pd.DataFrame:
    """Add zero rows for missing product/date pairs, then calculate day-over-day changes."""
    products = sorted(daily["product_id"].unique())
    dates = sorted(daily["date"].unique())
    full_index = pd.MultiIndex.from_product([products, dates], names=["product_id", "date"])
    complete = (
        daily.set_index(["product_id", "date"])
        .reindex(full_index)
        .reset_index()
        .sort_values(["product_id", "date"], ignore_index=True)
    )
    complete["sales_qty"] = complete["sales_qty"].fillna(0.0)
    complete["order_count"] = complete["order_count"].fillna(0.0)
    complete["profit"] = complete["profit"].fillna(0.0)
    complete["sheet"] = complete["date"].dt.strftime("%-m.%-d") if __import__("os").name != "nt" else complete["date"].apply(lambda d: f"{d.month}.{d.day}")
    complete["sales_change"] = complete.groupby("product_id")["sales_qty"].diff()
    complete["profit_change"] = complete.groupby("product_id")["profit"].diff()

    previous_sales = complete.groupby("product_id")["sales_qty"].shift()
    previous_profit = complete.groupby("product_id")["profit"].shift()
    complete["sales_change_pct"] = complete["sales_change"].div(previous_sales.where(previous_sales != 0))
    complete["profit_change_pct"] = complete["profit_change"].div(previous_profit.abs().where(previous_profit != 0))
    return complete


def build_summary(daily: pd.DataFrame, complete: pd.DataFrame) -> pd.DataFrame:
    total_dates = daily["date"].nunique()
    latest = complete.sort_values("date").groupby("product_id", as_index=False).tail(1)
    summary = (
        daily.groupby("product_id", as_index=False)
        .agg(
            total_sales=("sales_qty", "sum"),
            total_orders=("order_count", "sum"),
            total_profit=("profit", "sum"),
            active_days=("date", "nunique"),
            sku_count=("sku_count", "max"),
        )
        .merge(
            latest[
                ["product_id", "sales_qty", "order_count", "profit", "sales_change", "profit_change"]
            ].rename(
                columns={
                    "sales_qty": "latest_sales",
                    "order_count": "latest_orders",
                    "profit": "latest_profit",
                    "sales_change": "latest_sales_change",
                    "profit_change": "latest_profit_change",
                }
            ),
            on="product_id",
            how="left",
        )
    )
    summary["avg_daily_sales"] = summary["total_sales"] / total_dates
    summary["avg_daily_orders"] = summary["total_orders"] / total_dates
    summary["avg_daily_profit"] = summary["total_profit"] / total_dates
    return summary.sort_values(["total_sales", "total_profit"], ascending=[False, False], ignore_index=True)


def validate_known_sample(daily: pd.DataFrame) -> dict[str, float]:
    sample = daily[(daily["sheet"] == "8.13") & (daily["product_id"] == "655468070181")]
    if len(sample) != 1:
        raise AssertionError("样例校验失败：8.13 中商品 655468070181 应恰好聚合为一行")
    sales = float(sample.iloc[0]["sales_qty"])
    orders = float(sample.iloc[0]["order_count"])
    profit = float(sample.iloc[0]["profit"])
    if abs(sales - 20.0) > 1e-9 or abs(orders - 9.0) > 1e-9 or abs(profit - 22.015) > 1e-9:
        raise AssertionError(
            "样例校验失败：期望销量=20、订单量=9、盈亏=22.015，"
            f"实际销量={sales}、订单量={orders}、盈亏={profit}"
        )
    return {"sales_qty": sales, "order_count": orders, "profit": profit}
