# -*- coding: utf-8 -*-
"""
千牛多店铺 SKU 成本抓取器 V2.5.4
V2.5.2 核心变化：
- 不再用 context.request 重放 asyncSold（淘宝会返回错误页）
- 改为“浏览器真实请求 + 路由改参”
- 浏览器自己发请求，脚本只把日期/pageNum改成目标值
- 自动点击“下一页”，捕获真实 Response
"""

import asyncio
import csv
import json
import re
from pathlib import Path
from datetime import datetime, timedelta
from urllib.parse import parse_qsl, urlencode

from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook

from sku_cost_utils import normalize_sku_spec

SHOP_NAME = ""
CDP_PORT = 0
PAGE_SIZE = 15
MAX_PAGES = 500

DEFAULT_ORDER_URL = (
    "https://myseller.taobao.com/home.htm/trade-platform/tp/sold"
)

ORDER_URL = DEFAULT_ORDER_URL

BASE_DIR = Path(__file__).resolve().parent
SHOPS_FILE = BASE_DIR / "shops.json"
DATA_DIR = BASE_DIR / "data" / SHOP_NAME
LOG_DIR = BASE_DIR / "logs" / SHOP_NAME
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

RUN_DATE = datetime.now()
DATE_TEXT = RUN_DATE.strftime("%Y%m%d")

OUTPUT_CSV = DATA_DIR / f"order_sku_{DATE_TEXT}.csv"
MISSING_CODE_CSV = DATA_DIR / f"缺少商家编码_{DATE_TEXT}.csv"
FAILED_ITEM_CSV = DATA_DIR / f"商品ID解析失败_{DATE_TEXT}.csv"
CACHE_FILE = DATA_DIR / "trade_snap_cache.json"
RAW_JSON = LOG_DIR / f"order_raw_{DATE_TEXT}.json"

ORDER_COST_DETAIL_CSV = DATA_DIR / f"order_sku_cost_{DATE_TEXT}.csv"
PRODUCT_COST_SUMMARY_CSV = DATA_DIR / f"product_cost_summary_{DATE_TEXT}.csv"
UNMATCHED_COST_CSV = DATA_DIR / f"SKU成本未匹配_{DATE_TEXT}.csv"

CONFIG_DIR = BASE_DIR / "config"
CONFIG_DIR.mkdir(parents=True, exist_ok=True)

SKU_COST_FILE = CONFIG_DIR / "sku_cost.xlsx"

SKU_COST_HEADERS = [
    "店铺",
    "商品ID",
    "商家编码",
    "SKU规格",
    "单件货价",
    "快递费",
    "备注",
    "首次发现日期",
    "最近成交日期",
]



# ============================================================
# V2.3 多店铺配置
# ============================================================

def safe_filename(name):
    return re.sub(
        r'[\\/:*?"<>|]',
        "_",
        str(name)
    ).strip()


def load_enabled_shops():
    if not SHOPS_FILE.exists():
        raise RuntimeError(
            f"找不到 shops.json：{SHOPS_FILE}"
        )

    with open(
        SHOPS_FILE,
        "r",
        encoding="utf-8-sig"
    ) as f:
        config = json.load(f)

    shops = []
    used_ports = set()

    for item in config.get("shops", []):
        if not item.get("enabled", True):
            continue

        name = str(
            item.get("name", "")
        ).strip()

        if not name:
            continue

        try:
            port = int(
                item.get("port")
            )
        except Exception:
            print(
                f"⚠ {name} 的 Chrome 端口无效，跳过"
            )
            continue

        if port in used_ports:
            print(
                f"⚠ {name} 的 Chrome 端口 {port} 重复，跳过"
            )
            continue

        used_ports.add(port)

        shops.append({
            "name": name,
            "safe_name": safe_filename(name),
            "port": port,
            "order_url": str(
                item.get(
                    "order_url",
                    DEFAULT_ORDER_URL
                )
            ).strip() or DEFAULT_ORDER_URL,
        })

    return shops


def configure_shop(shop):
    """
    只切换店铺相关路径和端口。
    订单页面抓取/监听/翻页逻辑完全沿用已经验证成功的 V1.7，
    不再改动。
    """

    global SHOP_NAME
    global CDP_PORT
    global ORDER_URL
    global DATA_DIR
    global LOG_DIR
    global OUTPUT_CSV
    global MISSING_CODE_CSV
    global FAILED_ITEM_CSV
    global CACHE_FILE
    global RAW_JSON
    global ORDER_COST_DETAIL_CSV
    global PRODUCT_COST_SUMMARY_CSV
    global UNMATCHED_COST_CSV

    SHOP_NAME = shop["name"]
    CDP_PORT = int(shop["port"])
    ORDER_URL = (
        shop.get("order_url")
        or DEFAULT_ORDER_URL
    )

    DATA_DIR = (
        BASE_DIR
        /
        "data"
        /
        shop["safe_name"]
    )

    LOG_DIR = (
        BASE_DIR
        /
        "logs"
        /
        shop["safe_name"]
    )

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    LOG_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    OUTPUT_CSV = (
        DATA_DIR
        /
        f"order_sku_{DATE_TEXT}.csv"
    )

    MISSING_CODE_CSV = (
        DATA_DIR
        /
        f"缺少商家编码_{DATE_TEXT}.csv"
    )

    FAILED_ITEM_CSV = (
        DATA_DIR
        /
        f"商品ID解析失败_{DATE_TEXT}.csv"
    )

    CACHE_FILE = (
        DATA_DIR
        /
        "trade_snap_cache.json"
    )

    RAW_JSON = (
        LOG_DIR
        /
        f"order_raw_{DATE_TEXT}.json"
    )

    ORDER_COST_DETAIL_CSV = (
        DATA_DIR
        /
        f"order_sku_cost_{DATE_TEXT}.csv"
    )

    PRODUCT_COST_SUMMARY_CSV = (
        DATA_DIR
        /
        f"product_cost_summary_{DATE_TEXT}.csv"
    )

    UNMATCHED_COST_CSV = (
        DATA_DIR
        /
        f"SKU成本未匹配_{DATE_TEXT}.csv"
    )



def clean_text(v):
    return " ".join(str(v or "").replace("\r", " ").replace("\n", " ").replace("\t", " ").split())


def safe_int(v, default=0):
    try:
        return int(str(v).strip())
    except Exception:
        return default


def normalize_url(u):
    u = clean_text(u)
    if u.startswith("//"):
        return "https:" + u
    if u.startswith("/"):
        return "https://trade.taobao.com" + u
    return u


def today_timestamp_range():
    now = datetime.now()
    begin = datetime(now.year, now.month, now.day)
    end = begin + timedelta(days=1)
    return int(begin.timestamp() * 1000), int(end.timestamp() * 1000)


def load_cache():
    if not CACHE_FILE.exists():
        return {}
    try:
        d = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        return d if isinstance(d, dict) else {}
    except Exception:
        return {}


def save_cache(cache):
    CACHE_FILE.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def parse_sku_text(info):
    out = []
    if not isinstance(info, dict):
        return ""
    for x in info.get("skuText") or []:
        if not isinstance(x, dict):
            continue
        n = clean_text(x.get("name"))
        v = clean_text(x.get("value"))
        if n and v:
            out.append(f"{n}={v}")
        elif v:
            out.append(v)
    return " | ".join(out)


def parse_merchant_code(info):
    if not isinstance(info, dict):
        return ""
    for x in info.get("extra") or []:
        if isinstance(x, dict) and "商家编码" in clean_text(x.get("name")):
            return clean_text(x.get("value"))
    return ""


def find_trade_snap(obj):
    found = []

    def walk(x):
        if isinstance(x, dict):
            for k, v in x.items():
                if isinstance(v, str) and (
                    "tradesnap" in v.lower()
                    or "tradesnap" in str(k).lower()
                ):
                    found.append(v)
                walk(v)
        elif isinstance(x, list):
            for y in x:
                walk(y)

    walk(obj)
    return normalize_url(found[0]) if found else ""


def extract_order_items(data):
    rows = []

    def walk(o):
        if isinstance(o, dict):
            info = o.get("itemInfo")
            if isinstance(info, dict) and "skuText" in info:
                rows.append({
                    "order_id": clean_text(o.get("id")),
                    "quantity": safe_int(o.get("quantity"), 0),
                    "sku_text": parse_sku_text(info),
                    "merchant_code": parse_merchant_code(info),
                    "trade_snap": find_trade_snap(o),
                })
            for v in o.values():
                walk(v)
        elif isinstance(o, list):
            for v in o:
                walk(v)

    walk(data)

    seen, out = set(), []
    for r in rows:
        k = (
            r["order_id"],
            r["sku_text"],
            r["merchant_code"],
            r["quantity"],
            r["trade_snap"],
        )
        if k not in seen:
            seen.add(k)
            out.append(r)
    return out


def parse_json_like_text(text):
    text = (text or "").strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        pass
    a, b = text.find("{"), text.rfind("}")
    if a >= 0 and b > a:
        try:
            return json.loads(text[a:b+1])
        except Exception:
            pass
    return None


def parse_item_id_from_html(html):
    patterns = [
        r'detail\.tmall\.com[^"\']*?item\.htm\?id=(\d+)',
        r'item\.taobao\.com[^"\']*?item\.htm\?id=(\d+)',
        r'item\.htm\?id=(\d+)',
        r'商品ID\s*[:：]?\s*(\d+)',
    ]
    for p in patterns:
        m = re.search(p, html or "", re.I | re.S)
        if m:
            return m.group(1)
    return ""


async def get_item_id_from_trade_snap(context, url, cache):
    if not url:
        return "", "无交易快照", False

    url = normalize_url(url)

    if cache.get(url):
        return str(cache[url]), "缓存", True

    p = None
    try:
        p = await asyncio.wait_for(context.new_page(), timeout=10)
        await asyncio.wait_for(
            p.goto(url, wait_until="domcontentloaded", timeout=20000),
            timeout=25,
        )

        iid = parse_item_id_from_html(
            await asyncio.wait_for(p.content(), timeout=5)
        )

        if not iid:
            try:
                iid = parse_item_id_from_html(
                    await asyncio.wait_for(
                        p.locator("body").inner_text(timeout=3000),
                        timeout=5,
                    )
                )
            except Exception:
                pass

        if iid:
            cache[url] = iid
            return iid, "成功", False

        return "", "快照未解析到商品ID", False

    except Exception as e:
        return "", f"快照访问失败:{str(e)[:100]}", False

    finally:
        if p:
            try:
                await asyncio.wait_for(p.close(), timeout=5)
            except Exception:
                pass


CSV_FIELDS = [
    "店铺", "订单号", "商品ID", "购买数量",
    "SKU规格", "商家编码", "交易快照", "解析状态"
]


def save_rows_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        w.writeheader()

        for r in rows:
            w.writerow({
                "店铺": SHOP_NAME,
                "订单号": r.get("order_id", ""),
                "商品ID": r.get("item_id", ""),
                "购买数量": r.get("quantity", 0),
                "SKU规格": r.get("sku_text", ""),
                "商家编码": r.get("merchant_code", ""),
                "交易快照": r.get("trade_snap", ""),
                "解析状态": r.get("status", ""),
            })


async def get_order_page(context):
    """
    V2.5.3：
    优先寻找真正的订单页，而不是任意 trade-platform 页面。

    之前退款页：
      /trade-platform/refund-list
    也会被误认为订单页，导致三家都等不到 asyncSold。
    """
    pages = list(context.pages)

    # 1. 优先找明确的新版已卖出订单页
    for p in pages:
        u = (p.url or "").lower()

        if (
            "myseller.taobao.com" in u
            and
            (
                "/trade-platform/tp/sold" in u
                or
                "list_sold_items" in u
            )
        ):
            return p

    # 2. 找任意 myseller 页作为导航载体
    for p in pages:
        u = (p.url or "").lower()

        if "myseller.taobao.com" in u:
            return p

    return await context.new_page()


async def ensure_order_page(page):
    """
    无论当前浏览器停在退款页、首页还是其它千牛页面，
    都强制确认最终进入当前店铺配置的 ORDER_URL。
    """
    current = (page.url or "").lower()

    is_order_page = (
        "/trade-platform/tp/sold" in current
        or
        "list_sold_items" in current
    )

    if not is_order_page:
        print(
            f"当前不是订单页，自动跳转：{page.url}"
        )

        await page.goto(
            ORDER_URL,
            wait_until="domcontentloaded",
            timeout=60000
        )

    # 等 SPA 完成初始渲染
    await page.wait_for_timeout(
        2500
    )

    final_url = (
        page.url
        or
        ""
    ).lower()

    if not (
        "/trade-platform/tp/sold" in final_url
        or
        "list_sold_items" in final_url
    ):
        raise RuntimeError(
            f"未能进入订单页，当前URL：{page.url}"
        )

    print(
        f"订单页已确认：{page.url}"
    )


async def click_query_button(page):
    """
    V2.5.2：
    主动触发订单查询，不再只依赖 reload 自动产生 asyncSold。

    扫描 Page + Frame，优先点击订单查询区常见的：
      查询 / 搜索 / 搜索订单
    """
    scopes = [page]

    try:
        scopes.extend(list(page.frames))
    except Exception:
        pass

    labels = [
        "查询",
        "搜索",
        "搜索订单",
    ]

    for scope in scopes:
        for label in labels:
            try:
                loc = scope.get_by_text(
                    label,
                    exact=True
                )

                count = await loc.count()

                for i in range(
                    count - 1,
                    -1,
                    -1
                ):
                    item = loc.nth(i)

                    try:
                        if not await item.is_visible():
                            continue

                        await item.click(
                            timeout=4000
                        )

                        await page.wait_for_timeout(
                            250
                        )

                        return True

                    except Exception:
                        continue

            except Exception:
                continue

    # 最后尝试常见按钮 selector
    for selector in [
        "button:has-text('查询')",
        "button:has-text('搜索')",
        "a:has-text('查询')",
        "a:has-text('搜索')",
    ]:
        try:
            loc = page.locator(selector)
            count = await loc.count()

            for i in range(count):
                item = loc.nth(i)

                try:
                    if await item.is_visible():
                        await item.click(timeout=4000)
                        await page.wait_for_timeout(250)
                        return True
                except Exception:
                    continue
        except Exception:
            continue

    return False

class BrowserDrivenOrderPager:
    def __init__(self, page):
        self.page = page
        self.target_page = 1
        self.begin_ms, self.end_ms = today_timestamp_range()
        self.response_future = None

    async def route_handler(self, route, request):
        try:
            if "asyncSold.htm" not in request.url:
                await route.continue_()
                return

            post_data = request.post_data or ""

            if "SoldQueryAction" not in post_data:
                await route.continue_()
                return

            params = dict(parse_qsl(post_data, keep_blank_values=True))

            params["pageNum"] = str(self.target_page)
            params["prePageNo"] = str(max(0, self.target_page - 1))
            params["pageSize"] = str(PAGE_SIZE)
            params["payDateBegin"] = str(self.begin_ms)
            params["payDateEnd"] = str(self.end_ms)
            params["action"] = "itemlist/SoldQueryAction"

            new_post = urlencode(params)

            await route.continue_(post_data=new_post)

        except Exception:
            try:
                await route.continue_()
            except Exception:
                pass

    async def response_handler(self, response):
        try:
            if self.response_future is None or self.response_future.done():
                return

            if "asyncSold.htm" not in response.url:
                return

            post_data = response.request.post_data or ""
            if "SoldQueryAction" not in post_data:
                return

            params = dict(parse_qsl(post_data, keep_blank_values=True))
            page_num = safe_int(params.get("pageNum"), 0)

            if page_num != self.target_page:
                return

            text = await response.text()
            data = parse_json_like_text(text)

            if data is None:
                # 浏览器真实请求也不是JSON时，保存原文
                debug = LOG_DIR / f"browser_asyncSold_page_{self.target_page}.txt"
                debug.write_text(text, encoding="utf-8", errors="ignore")
                self.response_future.set_exception(
                    RuntimeError(
                        f"第{self.target_page}页浏览器返回不是JSON，已保存：{debug}"
                    )
                )
                return

            self.response_future.set_result(data)

        except Exception as e:
            if self.response_future is not None and not self.response_future.done():
                self.response_future.set_exception(e)

    async def start(self):
        await self.page.route("**/asyncSold.htm*", self.route_handler)
        self.page.on("response", self.response_handler)

    async def stop(self):
        try:
            await self.page.unroute("**/asyncSold.htm*", self.route_handler)
        except Exception:
            pass
        try:
            self.page.remove_listener("response", self.response_handler)
        except Exception:
            pass

    async def trigger_first_page(self):
        """
        V2.5.2：
        第一页不再只依赖 reload。

        顺序：
        1. 直接点击【查询/搜索】触发真实 asyncSold
        2. 若未触发，reload 后等待
        3. reload 后再主动点一次查询
        """

        self.target_page = 1

        async def wait_response(seconds):
            try:
                return await asyncio.wait_for(
                    asyncio.shield(
                        self.response_future
                    ),
                    timeout=seconds
                )
            except asyncio.TimeoutError:
                return None

        # 方案1：当前订单页直接查询
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        clicked = await click_query_button(
            self.page
        )

        if clicked:
            result = await wait_response(15)

            if result is not None:
                return result

        # 方案2：刷新订单页
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        try:
            await self.page.reload(
                wait_until="domcontentloaded",
                timeout=60000
            )
        except Exception:
            pass

        result = await wait_response(20)

        if result is not None:
            return result

        # 方案3：刷新后再次主动查询
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        clicked = await click_query_button(
            self.page
        )

        if clicked:
            result = await wait_response(15)

            if result is not None:
                return result

        raise RuntimeError(
            "订单页已尝试【直接查询 + 刷新 + 再查询】，"
            "仍未捕获第一页 asyncSold 请求"
        )

    async def trigger_next_page(self, page_num):
        """
        V2.5 修复：
        不再依赖“第5页/下一页”按钮是否出现在 DOM 中。

        原理：
        BrowserDrivenOrderPager 的 route_handler 会把任何真实
        asyncSold 查询请求的 pageNum 强制改成 self.target_page。

        所以后续页只需要触发“任意一次真实订单查询”即可，
        不必真的点击对应页码。

        优先顺序：
        1. 自动点击订单页【查询/搜索】按钮
        2. 找得到的话再点击【下一页】
        3. 最后 reload 页面

        route_handler 会把这些请求统一改成目标 pageNum。
        """

        self.target_page = page_num
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        # ----------------------------------------------------
        # 方案1：直接点击查询按钮
        # 最稳，不依赖分页页码是否可见
        # ----------------------------------------------------
        clicked = await click_query_button(
            self.page
        )

        if clicked:
            try:
                return await asyncio.wait_for(
                    self.response_future,
                    timeout=30
                )
            except asyncio.TimeoutError:
                pass

        # ----------------------------------------------------
        # 方案2：点击下一页
        # ----------------------------------------------------
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        clicked_next = False

        for sel in [
            "a:has-text('下一页')",
            "button:has-text('下一页')",
            "text=下一页",
        ]:
            try:
                loc = self.page.locator(sel)
                count = await loc.count()

                for i in range(
                    count - 1,
                    -1,
                    -1
                ):
                    item = loc.nth(i)

                    try:
                        if await item.is_visible():
                            await item.click(
                                timeout=5000
                            )
                            clicked_next = True
                            break
                    except Exception:
                        continue

                if clicked_next:
                    break

            except Exception:
                pass

        if clicked_next:
            try:
                return await asyncio.wait_for(
                    self.response_future,
                    timeout=30
                )
            except asyncio.TimeoutError:
                pass

        # ----------------------------------------------------
        # 方案3：reload 触发真实查询
        # ----------------------------------------------------
        self.response_future = (
            asyncio.get_running_loop()
            .create_future()
        )

        try:
            await self.page.reload(
                wait_until="domcontentloaded",
                timeout=60000
            )
        except Exception:
            pass

        try:
            return await asyncio.wait_for(
                self.response_future,
                timeout=30
            )
        except asyncio.TimeoutError:
            raise RuntimeError(
                f"无法触发第{page_num}页 asyncSold 请求"
            )


async def crawl_all_pages(page):
    pager = BrowserDrivenOrderPager(page)
    await pager.start()

    raw_pages = []
    all_rows = []
    seen = set()

    try:
        print("抓取第 1 页...", end="", flush=True)
        data = await pager.trigger_first_page()
        raw_pages.append(data)

        items = extract_order_items(data)
        for r in items:
            k = (
                r["order_id"], r["sku_text"], r["merchant_code"],
                r["quantity"], r["trade_snap"]
            )
            if k not in seen:
                seen.add(k)
                all_rows.append(r)

        print(f" {len(items)} 条，新增 {len(items)} 条")

        # V2.5：SKU行数不能用来判断是否为最后一页。
        if not items:
            return raw_pages, all_rows

        page_num = 2

        while page_num <= MAX_PAGES:
            print(f"抓取第 {page_num} 页...", end="", flush=True)

            data = await pager.trigger_next_page(page_num)
            raw_pages.append(data)

            items = extract_order_items(data)
            new_count = 0

            for r in items:
                k = (
                    r["order_id"], r["sku_text"], r["merchant_code"],
                    r["quantity"], r["trade_snap"]
                )
                if k not in seen:
                    seen.add(k)
                    all_rows.append(r)
                    new_count += 1

            print(f" {len(items)} 条，新增 {new_count} 条")

            # V2.5：
            # items 是“订单商品/SKU行数”，不是分页订单数。
            # 一个订单可能有多个 SKU，所以不能用 len(items) < PAGE_SIZE
            # 判断最后一页。
            #
            # 真正可靠的停止条件：
            # - 本页完全没有订单商品
            # - 或者本页没有任何新增记录（接口开始重复上一页/已越界）
            if not items:
                print("本页无订单商品，停止分页。")
                break

            if new_count == 0:
                print("⚠ 本页没有新增订单，判断已到最后一页，停止分页。")
                break

            page_num += 1
            await page.wait_for_timeout(200)

        return raw_pages, all_rows

    finally:
        await pager.stop()


# ============================================================
# SKU 成本配置
# ============================================================

def sku_match_key(row):
    """
    成本配置唯一键：
    1. 有商家编码：店铺 + 商品ID + 商家编码
    2. 无商家编码：店铺 + 商品ID + SKU规格

    这样即使部分 SKU 暂时没有商家编码，也能先进入成本表。
    后续补了商家编码后，程序会新增一条新的正式编码记录，
    不会覆盖旧记录。
    """
    shop = clean_text(row.get("店铺") or SHOP_NAME)
    item_id = clean_text(row.get("商品ID") or row.get("item_id"))
    code = clean_text(row.get("商家编码") or row.get("merchant_code"))
    sku = normalize_sku_spec(row.get("SKU规格") or row.get("sku_text"))

    if code:
        return ("code", shop, item_id, code)

    return ("sku", shop, item_id, sku)


def ensure_sku_cost_workbook(order_rows):
    """
    第一次运行：
      自动创建 config/sku_cost.xlsx

    以后运行：
      自动追加新 SKU
      不覆盖已填写的单件货价/快递费/备注
      更新最近成交日期
    """

    today_text = RUN_DATE.strftime("%Y-%m-%d")

    # --------------------------------------------------------
    # 读取旧配置
    # --------------------------------------------------------
    old_rows = []
    existing_by_key = {}

    if SKU_COST_FILE.exists():
        wb = load_workbook(SKU_COST_FILE)
        ws = wb.active

        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx

        # 如果旧表结构不完整，补齐表头
        for header in SKU_COST_HEADERS:
            if header not in headers:
                col = ws.max_column + 1
                ws.cell(row=1, column=col, value=header)
                headers[header] = col

        for r in range(2, ws.max_row + 1):
            row = {
                h: ws.cell(r, headers[h]).value
                for h in SKU_COST_HEADERS
            }
            key = sku_match_key(row)
            existing_by_key[key] = r
            old_rows.append(row)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "SKU成本配置"

        for col_idx, header in enumerate(SKU_COST_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        headers = {
            h: i + 1
            for i, h in enumerate(SKU_COST_HEADERS)
        }

        # 简单冻结表头
        ws.freeze_panes = "A2"

        # 列宽
        widths = {
            "A": 12,
            "B": 18,
            "C": 28,
            "D": 45,
            "E": 12,
            "F": 10,
            "G": 24,
            "H": 14,
            "I": 14,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    # --------------------------------------------------------
    # 从今天订单中生成不重复 SKU
    # --------------------------------------------------------
    unique = {}
    for row in order_rows:
        item_id = clean_text(row.get("item_id"))
        sku_text = clean_text(row.get("sku_text"))
        merchant_code = clean_text(row.get("merchant_code"))

        if not item_id:
            continue

        data = {
            "店铺": SHOP_NAME,
            "商品ID": item_id,
            "商家编码": merchant_code,
            "SKU规格": sku_text,
            "单件货价": None,
            "快递费": None,
            "备注": "",
            "首次发现日期": today_text,
            "最近成交日期": today_text,
        }

        key = sku_match_key(data)
        unique[key] = data

    added = 0
    updated = 0

    for key, data in unique.items():
        if key in existing_by_key:
            row_num = existing_by_key[key]

            # 只更新最近成交日期，以及缺失的商品/规格信息。
            ws.cell(
                row=row_num,
                column=headers["最近成交日期"],
                value=today_text
            )

            # 如果历史记录里商家编码为空，而当前有编码，理论上会变成新key，
            # 所以这里主要补充规格文本。
            if not ws.cell(row_num, headers["SKU规格"]).value and data["SKU规格"]:
                ws.cell(
                    row=row_num,
                    column=headers["SKU规格"],
                    value=data["SKU规格"]
                )

            updated += 1
            continue

        row_num = ws.max_row + 1

        for header in SKU_COST_HEADERS:
            ws.cell(
                row=row_num,
                column=headers[header],
                value=data.get(header)
            )

        existing_by_key[key] = row_num
        added += 1

    try:
        wb.save(
            SKU_COST_FILE
        )

    except PermissionError:
        locked_copy = (
            CONFIG_DIR
            /
            (
                "sku_cost_待合并_"
                +
                safe_filename(SHOP_NAME)
                +
                "_"
                +
                datetime.now().strftime(
                    "%Y%m%d_%H%M%S"
                )
                +
                ".xlsx"
            )
        )

        wb.save(
            locked_copy
        )

        print()
        print(
            f"⚠ sku_cost.xlsx 正在被 Excel 占用：{SKU_COST_FILE}"
        )
        print(
            f"   新SKU暂存：{locked_copy}"
        )
        print(
            "   关闭 Excel 后再运行一次即可自动追加。"
        )

    return added, updated, len(unique)


def sku_cost_summary(order_rows):
    """
    读取成本表并检查今天订单：
    - 已配置成本 SKU 数
    - 未配置成本 SKU 数
    - 缺商家编码 SKU 数
    """

    if not SKU_COST_FILE.exists():
        return {
            "configured": 0,
            "unconfigured": 0,
            "missing_code": 0,
        }

    wb = load_workbook(SKU_COST_FILE, data_only=True)
    ws = wb.active

    headers = {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws[1], start=1)
        if cell.value
    }

    cost_by_key = {}

    for r in range(2, ws.max_row + 1):
        row = {
            h: ws.cell(r, headers[h]).value
            if h in headers
            else None
            for h in SKU_COST_HEADERS
        }

        key = sku_match_key(row)
        cost_by_key[key] = row

    configured = 0
    unconfigured = 0
    missing_code = 0

    seen = set()

    for row in order_rows:
        if not row.get("item_id"):
            continue

        data = {
            "店铺": SHOP_NAME,
            "商品ID": row.get("item_id"),
            "商家编码": row.get("merchant_code"),
            "SKU规格": row.get("sku_text"),
        }

        key = sku_match_key(data)

        if key in seen:
            continue

        seen.add(key)

        if not clean_text(row.get("merchant_code")):
            missing_code += 1

        config_row = cost_by_key.get(key)

        if not config_row:
            unconfigured += 1
            continue

        price = config_row.get("单件货价")

        try:
            price = float(price)
            if price >= 0:
                configured += 1
            else:
                unconfigured += 1
        except Exception:
            unconfigured += 1

    return {
        "configured": configured,
        "unconfigured": unconfigured,
        "missing_code": missing_code,
    }



# ============================================================
# V1.7 SKU 成本匹配与汇总
# ============================================================

def safe_float(value, default=None):
    if value is None:
        return default

    try:
        text = str(value).strip()
        if not text:
            return default

        text = (
            text
            .replace("¥", "")
            .replace("￥", "")
            .replace(",", "")
        )

        return float(text)

    except Exception:
        return default


def load_sku_cost_maps():
    """
    从 config/sku_cost.xlsx 读取成本。

    匹配优先级：
    1. 店铺 + 商品ID + 商家编码
    2. 店铺 + 商品ID + SKU规格

    注意：
    商家编码不会单独作为唯一键，避免不同商品复用同一编码时串成本。
    """

    if not SKU_COST_FILE.exists():
        raise RuntimeError(
            f"SKU成本配置表不存在：{SKU_COST_FILE}"
        )

    wb = load_workbook(
        SKU_COST_FILE,
        data_only=True
    )

    ws = wb.active

    headers = {
        clean_text(cell.value): idx
        for idx, cell in enumerate(ws[1], start=1)
        if clean_text(cell.value)
    }

    required = [
        "店铺",
        "商品ID",
        "商家编码",
        "SKU规格",
        "单件货价",
        "快递费",
    ]

    missing_headers = [
        x
        for x in required
        if x not in headers
    ]

    if missing_headers:
        raise RuntimeError(
            "SKU成本表缺少字段："
            +
            "、".join(missing_headers)
        )

    code_map = {}
    sku_map = {}

    duplicate_code_keys = 0
    duplicate_sku_keys = 0

    for row_num in range(2, ws.max_row + 1):

        shop = clean_text(
            ws.cell(
                row_num,
                headers["店铺"]
            ).value
        )

        item_id = clean_text(
            ws.cell(
                row_num,
                headers["商品ID"]
            ).value
        )

        merchant_code = clean_text(
            ws.cell(
                row_num,
                headers["商家编码"]
            ).value
        )

        sku_text_value = clean_text(
            ws.cell(
                row_num,
                headers["SKU规格"]
            ).value
        )

        unit_cost = safe_float(
            ws.cell(
                row_num,
                headers["单件货价"]
            ).value
        )

        shipping_fee = safe_float(
            ws.cell(
                row_num,
                headers["快递费"]
            ).value
        )

        config = {
            "row_num": row_num,
            "shop": shop,
            "item_id": item_id,
            "merchant_code": merchant_code,
            "sku_text": sku_text_value,
            "unit_cost": unit_cost,
            "shipping_fee": shipping_fee,
        }

        if not item_id:
            continue

        if merchant_code:
            key = (
                shop,
                item_id,
                merchant_code
            )

            if key in code_map:
                duplicate_code_keys += 1

            code_map[key] = config

        if sku_text_value:
            key = (
                shop,
                item_id,
                normalize_sku_spec(sku_text_value)
            )

            if key in sku_map:
                duplicate_sku_keys += 1

            sku_map[key] = config

    return {
        "code_map": code_map,
        "sku_map": sku_map,
        "duplicate_code_keys": duplicate_code_keys,
        "duplicate_sku_keys": duplicate_sku_keys,
    }


def match_sku_cost(row, maps):
    """
    先按商品ID+商家编码匹配。
    编码不存在/未匹配时，退回商品ID+SKU规格。
    """

    item_id = clean_text(
        row.get("item_id")
    )

    merchant_code = clean_text(
        row.get("merchant_code")
    )

    sku_text_value = clean_text(
        row.get("sku_text")
    )

    if not item_id:
        return None, "商品ID缺失"

    if merchant_code:
        key = (
            SHOP_NAME,
            item_id,
            merchant_code
        )

        config = maps[
            "code_map"
        ].get(key)

        if config is not None:
            return (
                config,
                "商品ID+商家编码"
            )

    if sku_text_value:
        key = (
            SHOP_NAME,
            item_id,
            normalize_sku_spec(sku_text_value)
        )

        config = maps[
            "sku_map"
        ].get(key)

        if config is not None:
            return (
                config,
                "商品ID+SKU规格"
            )

    return None, "未匹配成本配置"


def calculate_order_sku_costs(final_rows):
    """
    计算每条订单商品的：
    - 单件货价
    - 货品成本

    快递费规则：
    - 同一个订单号只计算一次快递费
    - 若同一订单内 SKU 配置的快递费不同，取最大值作为该订单快递费
    - 订单快递费按各商品货品成本占比分摊
    - 若货品成本无法作为分母，则按购买数量占比分摊

    这样可避免一个订单买多个SKU时重复扣多次快递费。
    """

    maps = load_sku_cost_maps()

    detail_rows = []
    unmatched_rows = []

    # --------------------------------------------------------
    # 先匹配货价 / 快递费
    # --------------------------------------------------------
    for source in final_rows:

        row = dict(source)

        config, match_type = (
            match_sku_cost(
                row,
                maps
            )
        )

        quantity = max(
            0,
            safe_int(
                row.get("quantity"),
                0
            )
        )

        row[
            "cost_match_type"
        ] = match_type

        row[
            "unit_cost"
        ] = None

        row[
            "shipping_fee_config"
        ] = None

        row[
            "merchandise_cost"
        ] = None

        row[
            "shipping_cost_allocated"
        ] = 0.0

        row[
            "total_cost"
        ] = None

        row[
            "cost_status"
        ] = ""

        if config is None:

            row[
                "cost_status"
            ] = "未匹配SKU成本"

            unmatched_rows.append(
                row
            )

            detail_rows.append(
                row
            )

            continue

        unit_cost = config.get(
            "unit_cost"
        )

        shipping_fee = config.get(
            "shipping_fee"
        )

        row[
            "unit_cost"
        ] = unit_cost

        row[
            "shipping_fee_config"
        ] = shipping_fee

        if unit_cost is None:

            row[
                "cost_status"
            ] = "单件货价未填写"

            unmatched_rows.append(
                row
            )

            detail_rows.append(
                row
            )

            continue

        row[
            "merchandise_cost"
        ] = (
            quantity
            *
            unit_cost
        )

        if shipping_fee is None:

            row[
                "cost_status"
            ] = "快递费未填写"

            unmatched_rows.append(
                row
            )

        else:

            row[
                "cost_status"
            ] = "成本已匹配"

        detail_rows.append(
            row
        )

    # --------------------------------------------------------
    # 同订单只算一次快递费
    # --------------------------------------------------------
    orders = {}

    for row in detail_rows:

        order_id = clean_text(
            row.get(
                "order_id"
            )
        )

        if not order_id:
            # 无订单号时无法去重，单独作为一组
            order_id = (
                "__NO_ORDER__"
                +
                str(
                    len(orders) + 1
                )
            )

        orders.setdefault(
            order_id,
            []
        ).append(
            row
        )

    for order_id, rows in orders.items():

        shipping_values = [
            x.get(
                "shipping_fee_config"
            )
            for x in rows
            if x.get(
                "shipping_fee_config"
            ) is not None
        ]

        if shipping_values:

            order_shipping = max(
                shipping_values
            )

        else:

            order_shipping = 0.0

        unique_shipping = sorted(
            {
                round(
                    float(x),
                    6
                )
                for x in shipping_values
            }
        )

        shipping_note = ""

        if len(
            unique_shipping
        ) > 1:

            shipping_note = (
                "同订单存在多个快递费配置，"
                "本单取最大值"
            )

        positive_cost_rows = [
            x
            for x in rows
            if x.get(
                "merchandise_cost"
            ) is not None
        ]

        merchandise_total = sum(
            float(
                x.get(
                    "merchandise_cost"
                )
                or
                0
            )
            for x in positive_cost_rows
        )

        quantity_total = sum(
            max(
                0,
                safe_int(
                    x.get(
                        "quantity"
                    ),
                    0
                )
            )
            for x in rows
        )

        for row in rows:

            row[
                "order_shipping_fee"
            ] = order_shipping

            row[
                "shipping_rule_note"
            ] = shipping_note

            allocated = 0.0

            if (
                order_shipping > 0
                and
                len(rows) == 1
            ):

                allocated = order_shipping

            elif (
                order_shipping > 0
                and
                merchandise_total > 0
                and
                row.get(
                    "merchandise_cost"
                ) is not None
            ):

                allocated = (
                    order_shipping
                    *
                    float(
                        row[
                            "merchandise_cost"
                        ]
                    )
                    /
                    merchandise_total
                )

            elif (
                order_shipping > 0
                and
                quantity_total > 0
            ):

                allocated = (
                    order_shipping
                    *
                    max(
                        0,
                        safe_int(
                            row.get(
                                "quantity"
                            ),
                            0
                        )
                    )
                    /
                    quantity_total
                )

            row[
                "shipping_cost_allocated"
            ] = round(
                allocated,
                4
            )

            if row.get(
                "merchandise_cost"
            ) is not None:

                row[
                    "total_cost"
                ] = round(
                    float(
                        row[
                            "merchandise_cost"
                        ]
                    )
                    +
                    row[
                        "shipping_cost_allocated"
                    ],
                    4
                )

    return (
        detail_rows,
        unmatched_rows,
        maps
    )


ORDER_COST_FIELDS = [
    "店铺",
    "订单号",
    "商品ID",
    "购买数量",
    "SKU规格",
    "商家编码",
    "匹配方式",
    "单件货价",
    "货品成本",
    "SKU配置快递费",
    "订单快递费",
    "分摊快递费",
    "货品+快递成本",
    "成本状态",
    "快递费处理说明",
    "交易快照",
]


def save_order_cost_detail(
    path,
    rows
):

    with open(
        path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=ORDER_COST_FIELDS
        )

        writer.writeheader()

        for row in rows:

            writer.writerow({
                "店铺":
                    SHOP_NAME,

                "订单号":
                    row.get(
                        "order_id",
                        ""
                    ),

                "商品ID":
                    row.get(
                        "item_id",
                        ""
                    ),

                "购买数量":
                    row.get(
                        "quantity",
                        0
                    ),

                "SKU规格":
                    row.get(
                        "sku_text",
                        ""
                    ),

                "商家编码":
                    row.get(
                        "merchant_code",
                        ""
                    ),

                "匹配方式":
                    row.get(
                        "cost_match_type",
                        ""
                    ),

                "单件货价":
                    (
                        ""
                        if row.get(
                            "unit_cost"
                        ) is None
                        else round(
                            float(
                                row[
                                    "unit_cost"
                                ]
                            ),
                            4
                        )
                    ),

                "货品成本":
                    (
                        ""
                        if row.get(
                            "merchandise_cost"
                        ) is None
                        else round(
                            float(
                                row[
                                    "merchandise_cost"
                                ]
                            ),
                            4
                        )
                    ),

                "SKU配置快递费":
                    (
                        ""
                        if row.get(
                            "shipping_fee_config"
                        ) is None
                        else round(
                            float(
                                row[
                                    "shipping_fee_config"
                                ]
                            ),
                            4
                        )
                    ),

                "订单快递费":
                    round(
                        float(
                            row.get(
                                "order_shipping_fee",
                                0
                            )
                            or
                            0
                        ),
                        4
                    ),

                "分摊快递费":
                    round(
                        float(
                            row.get(
                                "shipping_cost_allocated",
                                0
                            )
                            or
                            0
                        ),
                        4
                    ),

                "货品+快递成本":
                    (
                        ""
                        if row.get(
                            "total_cost"
                        ) is None
                        else round(
                            float(
                                row[
                                    "total_cost"
                                ]
                            ),
                            4
                        )
                    ),

                "成本状态":
                    row.get(
                        "cost_status",
                        ""
                    ),

                "快递费处理说明":
                    row.get(
                        "shipping_rule_note",
                        ""
                    ),

                "交易快照":
                    row.get(
                        "trade_snap",
                        ""
                    ),
            })


PRODUCT_COST_FIELDS = [
    "店铺",
    "商品ID",
    "订单数",
    "成交件数",
    "货品成本",
    "分摊快递费",
    "货品+快递总成本",
    "未匹配成本行数",
]


def build_product_cost_summary(
    detail_rows
):

    groups = {}

    for row in detail_rows:

        item_id = clean_text(
            row.get(
                "item_id"
            )
        )

        if not item_id:
            continue

        group = groups.setdefault(
            item_id,
            {
                "order_ids":
                    set(),

                "quantity":
                    0,

                "merchandise_cost":
                    0.0,

                "shipping_cost":
                    0.0,

                "unmatched":
                    0,
            }
        )

        order_id = clean_text(
            row.get(
                "order_id"
            )
        )

        if order_id:
            group[
                "order_ids"
            ].add(
                order_id
            )

        group[
            "quantity"
        ] += max(
            0,
            safe_int(
                row.get(
                    "quantity"
                ),
                0
            )
        )

        if row.get(
            "merchandise_cost"
        ) is not None:

            group[
                "merchandise_cost"
            ] += float(
                row[
                    "merchandise_cost"
                ]
            )

        group[
            "shipping_cost"
        ] += float(
            row.get(
                "shipping_cost_allocated",
                0
            )
            or
            0
        )

        if row.get(
            "cost_status"
        ) != "成本已匹配":

            group[
                "unmatched"
            ] += 1

    output = []

    for item_id, g in groups.items():

        merchandise = round(
            g[
                "merchandise_cost"
            ],
            4
        )

        shipping = round(
            g[
                "shipping_cost"
            ],
            4
        )

        output.append({
            "店铺":
                SHOP_NAME,

            "商品ID":
                item_id,

            "订单数":
                len(
                    g[
                        "order_ids"
                    ]
                ),

            "成交件数":
                g[
                    "quantity"
                ],

            "货品成本":
                merchandise,

            "分摊快递费":
                shipping,

            "货品+快递总成本":
                round(
                    merchandise
                    +
                    shipping,
                    4
                ),

            "未匹配成本行数":
                g[
                    "unmatched"
                ],
        })

    output.sort(
        key=lambda x: (
            -float(
                x[
                    "货品+快递总成本"
                ]
            ),
            x[
                "商品ID"
            ],
        )
    )

    return output


def save_product_cost_summary(
    path,
    rows
):

    with open(
        path,
        "w",
        newline="",
        encoding="utf-8-sig"
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=PRODUCT_COST_FIELDS
        )

        writer.writeheader()

        writer.writerows(
            rows
        )


def save_unmatched_cost_rows(
    path,
    rows
):
    save_order_cost_detail(
        path,
        rows
    )



def get_cdp_websocket_url(port):
    """
    从 Chrome /json/version 读取 browser websocket 地址。
    Playwright 用 http://port 连接超时时，可直接连接这个 ws 地址。
    """
    url = f"http://127.0.0.1:{port}/json/version"

    try:
        with urllib.request.urlopen(
            url,
            timeout=5
        ) as resp:
            data = json.loads(
                resp.read().decode(
                    "utf-8",
                    errors="ignore"
                )
            )

        return str(
            data.get(
                "webSocketDebuggerUrl",
                ""
            )
        ).strip()

    except Exception:
        return ""


async def connect_cdp_robust(pw, port):
    """
    V2.5.4：
    CDP连接偶发卡在 websocket connecting。
    改为快速失败 + 重试，而不是单次等180秒。

    顺序：
      1. http endpoint，20秒
      2. 读取 /json/version 后直接 ws，20秒
      3. 再重试一次 http，25秒
    """
    endpoint = f"http://127.0.0.1:{port}"
    errors = []

    for attempt in range(1, 4):
        print(
            f"连接 Chrome CDP（第{attempt}/3次）..."
        )

        try:
            if attempt == 2:
                ws_url = get_cdp_websocket_url(
                    port
                )

                if ws_url:
                    browser = await pw.chromium.connect_over_cdp(
                        ws_url,
                        timeout=20000
                    )
                else:
                    browser = await pw.chromium.connect_over_cdp(
                        endpoint,
                        timeout=20000
                    )

            else:
                browser = await pw.chromium.connect_over_cdp(
                    endpoint,
                    timeout=25000 if attempt == 3 else 20000
                )

            print("✓ Chrome CDP连接成功")
            return browser

        except Exception as e:
            errors.append(
                f"第{attempt}次：{type(e).__name__}: {e}"
            )

            print(
                f"⚠ CDP连接失败，第{attempt}次重试..."
            )

            await asyncio.sleep(
                2
            )

    raise RuntimeError(
        "Chrome调试端口可访问，但Playwright连续3次无法建立CDP连接。"
        f" 端口={port}。"
        " 这通常是该Chrome调试会话卡住，不是SKU接口问题。"
        " 请只关闭并重新启动该端口对应的Chrome后再运行。"
        " 连接记录：" + " | ".join(errors)
    )


async def run_current_shop():
    print("=" * 72)
    print("千牛多店铺 SKU 成本抓取器 V2.5.4")
    print("自动分页 + SKU成本匹配 + 商品成本汇总")
    print("=" * 72)
    print(f"店铺：{SHOP_NAME}")
    print(f"Chrome CDP：127.0.0.1:{CDP_PORT}")
    print(f"抓取日期：{RUN_DATE.strftime('%Y-%m-%d')}")
    print()

    cache = load_cache()

    async with async_playwright() as pw:
        browser = await connect_cdp_robust(
            pw,
            CDP_PORT
        )

        if not browser.contexts:
            raise RuntimeError("没有找到 Chrome Context")

        context = browser.contexts[0]
        page = await get_order_page(context)

        await ensure_order_page(page)

        print("开始通过浏览器真实请求抓取订单...\n")

        raw_pages, items = await crawl_all_pages(page)

        RAW_JSON.write_text(
            json.dumps(raw_pages, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

        print(f"\n✓ 共抓到订单商品：{len(items)} 条")

        if not items:
            print(f"原始JSON：{RAW_JSON}")
            return

        print("\n开始解析交易快照商品ID...\n")

        final_rows = []
        cache_hit = 0
        new_snap = 0

        for index, row in enumerate(items, start=1):
            try:
                iid, status, hit = await asyncio.wait_for(
                    get_item_id_from_trade_snap(
                        context,
                        row.get("trade_snap", ""),
                        cache
                    ),
                    timeout=45,
                )
            except asyncio.TimeoutError:
                iid, status, hit = "", "trade snap timeout", False

            row["item_id"] = iid
            row["status"] = status

            cache_hit += int(hit)
            new_snap += int(bool(iid) and not hit)

            print(
                f"[{index}/{len(items)}] "
                f"订单 {row.get('order_id')} | "
                f"商品ID {iid or '失败'} | "
                f"商家编码 {row.get('merchant_code') or '缺失'}"
            )

            final_rows.append(row)

            if index % 20 == 0:
                save_cache(cache)

        save_cache(cache)

        save_rows_csv(OUTPUT_CSV, final_rows)

        missing_rows = [
            x for x in final_rows
            if not x.get("merchant_code")
        ]

        failed_rows = [
            x for x in final_rows
            if not x.get("item_id")
        ]

        save_rows_csv(MISSING_CODE_CSV, missing_rows)
        save_rows_csv(FAILED_ITEM_CSV, failed_rows)

        # ----------------------------------------------------
        # 自动维护 SKU 成本配置表
        # ----------------------------------------------------
        added_sku, updated_sku, today_unique_sku = ensure_sku_cost_workbook(
            final_rows
        )

        cost_check = sku_cost_summary(
            final_rows
        )

        # ----------------------------------------------------
        # V1.7：计算 SKU 实际货品成本 + 订单级快递费
        # ----------------------------------------------------
        (
            order_cost_rows,
            unmatched_cost_rows,
            cost_maps

        ) = calculate_order_sku_costs(
            final_rows
        )

        product_cost_rows = build_product_cost_summary(
            order_cost_rows
        )

        save_order_cost_detail(
            ORDER_COST_DETAIL_CSV,
            order_cost_rows
        )

        save_product_cost_summary(
            PRODUCT_COST_SUMMARY_CSV,
            product_cost_rows
        )

        save_unmatched_cost_rows(
            UNMATCHED_COST_CSV,
            unmatched_cost_rows
        )

        total_merchandise_cost = sum(
            float(
                x.get(
                    "merchandise_cost",
                    0
                )
                or
                0
            )
            for x in order_cost_rows
        )

        # 订单快递费只统计一次：
        order_shipping_map = {}

        for x in order_cost_rows:

            order_id = clean_text(
                x.get(
                    "order_id"
                )
            )

            if not order_id:
                continue

            order_shipping_map[
                order_id
            ] = max(
                float(
                    order_shipping_map.get(
                        order_id,
                        0
                    )
                ),
                float(
                    x.get(
                        "order_shipping_fee",
                        0
                    )
                    or
                    0
                )
            )

        total_shipping_cost = sum(
            order_shipping_map.values()
        )

        print("\n" + "=" * 72)
        print(f"{SHOP_NAME} SKU订单抓取完成")
        print("=" * 72)
        print(f"订单商品数：       {len(final_rows)}")
        print(f"商品ID成功：       {len(final_rows) - len(failed_rows)}")
        print(f"商品ID失败：       {len(failed_rows)}")
        print(f"缺少商家编码：     {len(missing_rows)}")
        print(f"交易快照缓存命中： {cache_hit}")
        print(f"本次新解析快照：   {new_snap}")
        print("-" * 72)
        print(f"今日唯一SKU：       {today_unique_sku}")
        print(f"成本表新增SKU：     {added_sku}")
        print(f"成本表已有SKU：     {updated_sku}")
        print(f"已配置单件货价：   {cost_check['configured']}")
        print(f"未配置单件货价：   {cost_check['unconfigured']}")
        print(f"缺少商家编码SKU：  {cost_check['missing_code']}")
        print(f"SKU成本配置表：    {SKU_COST_FILE}")
        print("-" * 72)
        print(f"SKU货品成本合计：  ¥{total_merchandise_cost:.2f}")
        print(f"订单快递费合计：   ¥{total_shipping_cost:.2f}")
        print(
            f"货品+快递总成本：  "
            f"¥{(total_merchandise_cost + total_shipping_cost):.2f}"
        )
        print(f"成本未匹配行数：   {len(unmatched_cost_rows)}")
        print(
            f"重复编码配置键：   "
            f"{cost_maps['duplicate_code_keys']}"
        )
        print(
            f"重复规格配置键：   "
            f"{cost_maps['duplicate_sku_keys']}"
        )
        print(f"\n完整订单SKU：{OUTPUT_CSV}")
        print(f"订单SKU成本：{ORDER_COST_DETAIL_CSV}")
        print(f"商品成本汇总：{PRODUCT_COST_SUMMARY_CSV}")
        print(f"成本未匹配：{UNMATCHED_COST_CSV}")
        print(f"缺少商家编码：{MISSING_CODE_CSV}")
        print(f"商品ID解析失败：{FAILED_ITEM_CSV}")
        print(f"交易快照缓存：{CACHE_FILE}")
        print("=" * 72)


# ============================================================
# V2.3 多店铺总入口
# ============================================================

async def main():

    print()
    print("=" * 76)
    print("千牛多店铺 SKU 成本抓取器 V2.5.4")
    print("核心订单抓取逻辑 = 已验证成功的 V1.7 原版")
    print("=" * 76)

    shops = load_enabled_shops()

    print(
        f"启用店铺：{len(shops)} 家"
    )

    for i, shop in enumerate(
        shops,
        start=1
    ):
        print(
            f"  {i}. {shop['name']}  "
            f"CDP={shop['port']}  "
            f"订单页={shop['order_url']}"
        )

    if not shops:
        return

    print()
    print(
        "运行前建议关闭：config\\sku_cost.xlsx"
    )

    success = 0
    failed = 0
    failures = []

    for i, shop in enumerate(
        shops,
        start=1
    ):

        configure_shop(
            shop
        )

        print()
        print()
        print("#" * 76)
        print(
            f"# [{i}/{len(shops)}] "
            f"{SHOP_NAME} / CDP {CDP_PORT}"
        )
        print(
            f"# 订单页：{ORDER_URL}"
        )
        print("#" * 76)

        try:
            await run_current_shop()
            success += 1

        except Exception as e:
            failed += 1

            failures.append(
                (
                    SHOP_NAME,
                    f"{type(e).__name__}: {e}"
                )
            )

            print()
            print(
                f"❌ [{SHOP_NAME}] SKU抓取失败"
            )
            print(
                f"{type(e).__name__}: {e}"
            )

            # 一家失败继续下一家
            continue

    print()
    print("=" * 76)
    print("多店铺 SKU 成本抓取结束")
    print("=" * 76)
    print(
        f"成功：{success} 家"
    )
    print(
        f"失败：{failed} 家"
    )

    if failures:
        print()
        print("失败店铺：")

        for name, err in failures:
            print(
                f"  - {name}: {err}"
            )

    print()
    print("商品成本汇总文件：")

    for shop in shops:
        path = (
            BASE_DIR
            /
            "data"
            /
            shop["safe_name"]
            /
            f"product_cost_summary_{DATE_TEXT}.csv"
        )

        print(
            f"  {'✓' if path.exists() else '×'} "
            f"{shop['name']}：{path}"
        )

    print()
    print(
        f"统一SKU成本表：{SKU_COST_FILE}"
    )
    print("=" * 76)

    return {
        "success": success,
        "failed": failed,
        "success_shops": [
            shop["name"]
            for shop in shops
            if shop["name"] not in {
                name for name, _ in failures
            }
        ],
        "failures": failures,
    }


if __name__ == "__main__":

    try:
        asyncio.run(
            main()
        )

    except KeyboardInterrupt:
        print(
            "\n用户停止程序"
        )

    except Exception as e:
        print()
        print(
            f"程序异常：{type(e).__name__}: {e}"
        )

    input(
        "\n按 Enter 退出..."
    )
