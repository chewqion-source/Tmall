from __future__ import annotations

import hashlib
import json
import os
import tempfile
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Mapping

import pandas as pd

from data_loader import (
    DATE_SHEET_PATTERN,
    STORE_FILE_PATTERNS,
    SUPPORTED_EXTENSIONS,
    SheetData,
    _infer_year,
    _iter_sheets,
    load_product_daily,
)


MAX_UPLOAD_BYTES = 25 * 1024 * 1024
STORE_CANONICAL_BASENAMES = {
    "易丽洁": "2026年天猫日报表-易丽洁",
    "坐拥宁静": "2026年坐拥宁静日报表",
    "国货严选": "2026年国货严选日报表",
    "咖时光": "2026年咖时光日报表-天猫",
}

STORE_NAME_ALIASES = {
    "易丽洁": ["易丽洁", "易丽洁日报", "天猫日报表易丽洁"],
    "坐拥宁静": ["坐拥宁静", "坐拥宁静日报"],
    "国货严选": ["国货严选", "国货严选日报"],
    "咖时光": ["咖时光", "咖时光日报", "咖时光天猫"],
}


@dataclass(frozen=True)
class UploadResult:
    store: str
    original_name: str
    saved_name: str
    sha256: str
    bytes: int
    start_date: str
    end_date: str
    date_sheets: int
    products: int
    sales: float
    orders: float
    profit: float
    added_dates: int = 0
    updated_dates: int = 0
    skipped_dates: int = 0


def _validate_upload(store: str, original_name: str, content: bytes) -> str:
    if store not in STORE_CANONICAL_BASENAMES:
        raise ValueError(f"未知店铺：{store}")
    suffix = Path(original_name).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"{store}：只支持 {', '.join(SUPPORTED_EXTENSIONS)}")
    if not content:
        raise ValueError(f"{store}：上传文件为空")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(f"{store}：文件超过 25MB")
    _validate_store_name_match(store, original_name)
    return suffix


def _normalize_filename_text(value: str) -> str:
    removable = ["2026", "2025", "天猫", "日报表", "日报", "报表", "财务", "数据", "日表"]
    text = Path(value).stem.lower()
    for word in removable:
        text = text.replace(word.lower(), "")
    return "".join(ch for ch in text if ch.isalnum() or "\u4e00" <= ch <= "\u9fff")


def _store_match_score(filename_text: str, aliases: list[str]) -> float:
    if not filename_text:
        return 0
    scores = []
    for alias in aliases:
        alias_text = _normalize_filename_text(alias)
        if alias_text and alias_text in filename_text:
            scores.append(1.0)
        else:
            scores.append(SequenceMatcher(None, filename_text, alias_text).ratio())
    return max(scores, default=0)


def _validate_store_name_match(store: str, original_name: str) -> None:
    filename_text = _normalize_filename_text(original_name)
    scores = {
        candidate: _store_match_score(filename_text, aliases)
        for candidate, aliases in STORE_NAME_ALIASES.items()
    }
    best_store, best_score = max(scores.items(), key=lambda item: item[1])
    selected_score = scores.get(store, 0)
    if best_score >= 0.62 and best_store != store:
        raise ValueError(
            f"{store}：文件名「{Path(original_name).name}」更像是「{best_store}」的报表，请确认店铺选择。"
        )
    if selected_score < 0.42:
        raise ValueError(
            f"{store}：文件名「{Path(original_name).name}」没有明显匹配「{store}」，请确认没有选错店铺。"
        )


def _existing_store_files(data_dir: Path, store: str) -> list[Path]:
    return sorted(
        {
            path
            for pattern in STORE_FILE_PATTERNS[store]
            for path in data_dir.glob(pattern)
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
        }
    )


def _sheet_date(sheet_name: str, year: int) -> pd.Timestamp | None:
    match = DATE_SHEET_PATTERN.fullmatch(sheet_name.strip())
    if not match:
        return None
    return pd.Timestamp(year, int(match.group(1)), int(match.group(2)))


def _date_sheets(path: Path) -> dict[pd.Timestamp, SheetData]:
    year = _infer_year(path)
    sheets: dict[pd.Timestamp, SheetData] = {}
    for sheet in _iter_sheets(path):
        sheet_date = _sheet_date(sheet.name, year)
        if sheet_date is not None:
            sheets[sheet_date] = sheet
    return sheets


def _daily_fingerprints(path: Path) -> dict[pd.Timestamp, str]:
    daily = load_product_daily(path)
    fingerprints: dict[pd.Timestamp, str] = {}
    compare_columns = ["product_id", "sales_qty", "order_count", "profit", "sku_count", "skus"]
    for sheet_date, group in daily.groupby("date"):
        comparable = group[compare_columns].sort_values("product_id", ignore_index=True)
        payload = comparable.to_json(orient="records", force_ascii=False, double_precision=6)
        fingerprints[pd.Timestamp(sheet_date)] = hashlib.sha256(payload.encode("utf-8")).hexdigest()
    return fingerprints


def _write_merged_workbook(sheets_by_date: Mapping[pd.Timestamp, SheetData], target_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_date in sorted(sheets_by_date):
        source = sheets_by_date[sheet_date]
        sheet = workbook.create_sheet(source.name[:31])
        for row_index, row in enumerate(source.rows, start=1):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
        for row_start, row_end, col_start, col_end in source.merged_ranges:
            if row_end > row_start + 1 or col_end > col_start + 1:
                sheet.merge_cells(
                    start_row=row_start + 1,
                    start_column=col_start + 1,
                    end_row=row_end,
                    end_column=col_end,
                )

    temporary_path = target_path.with_name(f".{target_path.name}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, target_path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)


def _merge_uploaded_workbook(staged_path: Path, target_path: Path) -> tuple[int, int, int]:
    incoming_sheets = _date_sheets(staged_path)
    incoming_fingerprints = _daily_fingerprints(staged_path)
    if not incoming_sheets:
        raise ValueError("上传文件没有可合并的日期 Sheet")

    existing_path = target_path if target_path.exists() else None
    if existing_path is None:
        existing_candidates = _existing_store_files(target_path.parent, _store_from_target(target_path))
        existing_path = (
            max(existing_candidates, key=lambda path: (path.stat().st_mtime_ns, path.name))
            if existing_candidates
            else None
        )

    existing_sheets: dict[pd.Timestamp, SheetData] = {}
    existing_fingerprints: dict[pd.Timestamp, str] = {}
    if existing_path:
        try:
            existing_sheets = _date_sheets(existing_path)
            existing_fingerprints = _daily_fingerprints(existing_path)
        except Exception:
            existing_sheets = {}
            existing_fingerprints = {}

    merged_sheets = dict(existing_sheets)
    added = updated = skipped = 0
    for sheet_date, sheet in incoming_sheets.items():
        if sheet_date not in existing_sheets:
            added += 1
            merged_sheets[sheet_date] = sheet
        elif existing_fingerprints.get(sheet_date) != incoming_fingerprints.get(sheet_date):
            updated += 1
            merged_sheets[sheet_date] = sheet
        else:
            skipped += 1

    _write_merged_workbook(merged_sheets, target_path)
    return added, updated, skipped


def _store_from_target(target_path: Path) -> str:
    stem = target_path.stem
    for store, basename in STORE_CANONICAL_BASENAMES.items():
        if stem.startswith(basename):
            return store
    raise ValueError(f"无法识别目标文件所属店铺：{target_path.name}")


def install_uploaded_workbooks(
    uploads: Mapping[str, tuple[str, bytes]], data_dir: str | Path
) -> list[UploadResult]:
    """Validate every workbook, then merge uploaded date sheets atomically.

    The caller supplies an explicit store for each file, so finance users do not
    have to preserve a particular filename. File contents are parsed before the
    current workbook is changed. Existing historical dates are preserved unless
    the uploaded date has different aggregate data, in which case that date is
    overwritten.
    """
    if not uploads:
        raise ValueError("请至少选择一家店铺的日报")

    target_dir = Path(data_dir).expanduser().resolve()
    target_dir.mkdir(parents=True, exist_ok=True)
    incoming_dir = target_dir / ".incoming"
    incoming_dir.mkdir(exist_ok=True)
    prepared: list[tuple[str, str, Path, Path, UploadResult]] = []
    staged_paths: list[Path] = []

    try:
        for store, (original_name, content) in uploads.items():
            suffix = _validate_upload(store, original_name, content)
            canonical_name = f"{STORE_CANONICAL_BASENAMES[store]}.xlsx"
            target_path = target_dir / canonical_name
            with tempfile.NamedTemporaryFile(
                mode="wb",
                dir=incoming_dir,
                prefix=f"{STORE_CANONICAL_BASENAMES[store]}-",
                suffix=suffix,
                delete=False,
            ) as temporary_file:
                temporary_file.write(content)
                staged_path = Path(temporary_file.name)
                staged_paths.append(staged_path)

            daily = load_product_daily(staged_path)
            result = UploadResult(
                store=store,
                original_name=Path(original_name).name,
                saved_name=canonical_name,
                sha256=hashlib.sha256(content).hexdigest(),
                bytes=len(content),
                start_date=daily["date"].min().strftime("%Y-%m-%d"),
                end_date=daily["date"].max().strftime("%Y-%m-%d"),
                date_sheets=int(daily["date"].nunique()),
                products=int(daily["product_id"].nunique()),
                sales=float(daily["sales_qty"].sum()),
                orders=float(daily["order_count"].sum()),
                profit=float(daily["profit"].sum()),
            )
            prepared.append((store, original_name, staged_path, target_path, result))

        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        results: list[UploadResult] = []
        for store, _original_name, staged_path, target_path, result in prepared:
            archive_dir = target_dir / "archive" / store
            archive_dir.mkdir(parents=True, exist_ok=True)
            existing_files = _existing_store_files(target_dir, store)
            for existing in existing_files:
                archive_name = f"{timestamp}-{uuid.uuid4().hex[:8]}-{existing.name}"
                archive_copy = archive_dir / archive_name
                archive_copy.write_bytes(existing.read_bytes())
            added, updated, skipped = _merge_uploaded_workbook(staged_path, target_path)
            if os.name != "nt":
                target_path.chmod(0o640)
            result = UploadResult(
                **{
                    **asdict(result),
                    "saved_name": target_path.name,
                    "added_dates": added,
                    "updated_dates": updated,
                    "skipped_dates": skipped,
                }
            )
            results.append(result)

        history_path = target_dir / "upload-history.jsonl"
        history_record = {
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "files": [asdict(result) for result in results],
        }
        with history_path.open("a", encoding="utf-8") as history_file:
            history_file.write(json.dumps(history_record, ensure_ascii=False) + "\n")
        return results
    finally:
        for staged_path in staged_paths:
            staged_path.unlink(missing_ok=True)
